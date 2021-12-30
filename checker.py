from io import DEFAULT_BUFFER_SIZE
import os
from os import path
import shutil
import json
import tkinter
from tkinter import filedialog
import pandas as pd
import json
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException
import time

root = tkinter.Tk()
root.withdraw()

# Open xlsx file
open_sheet = path.exists("cache/opened_sheet.json")
if open_sheet == True :
  opened_sheet_file_path = "cache/opened_sheet.json"
  json_file = open(opened_sheet_file_path)
  data = json.load(json_file)
  xlsx_sheet_check = path.exists(data ['xlsx_file_path'])
  if xlsx_sheet_check == True :
    xlsx_file_path = data ['xlsx_file_path']
  else :
    shutil.rmtree('cache', ignore_errors=True)
    xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
    cache_path = os.path.join(str(os.getcwd()), "cache")
    dictionary = {"xlsx_file_path" : xlsx_file_path}
    json_object = json.dumps(dictionary, indent = 1)
    with open("cache/opened_sheet.json", "w") as outfile:
      outfile.write(json_object)
else :
  xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
  cache_path = os.path.join(str(os.getcwd()), "cache")
  os.mkdir(cache_path)
  dictionary = {"xlsx_file_path" : xlsx_file_path}
  json_object = json.dumps(dictionary, indent = 1)
  with open("cache/opened_sheet.json", "w") as outfile:
    outfile.write(json_object)

# read imported xlsx file path using pandas
input_workbook = pd.read_excel(xlsx_file_path, sheet_name = 'Sheet1', usecols = 'A', dtype=str)
input_workbook.head()

# read total number of rows present in xlsx
number_of_rows = len(input_workbook.index)
print('Total Cards = ',number_of_rows)

input_workbook_mobile_number = input_workbook['MOBILE NUMBER'].values.tolist()

def start_link():
    driver.get("https://gp-giftcard.novopay.in/GeoPay/Login")

def demo():
  global balance, cardno, expiry, cvv, kitno
  driver.find_element_by_id(cards_list[n]).click()
  time.sleep(3)
  driver.find_element_by_xpath('//*[@id="View-Details"]').click()
  driver.find_element_by_xpath('//*[@id="auth_value"]').send_keys("2021")
  driver.find_element_by_xpath('//*[@id="im_footer"]/div/button').click()
  time.sleep(1)
  cardno = driver.find_element_by_id('card_number_fill').text
  expiry = driver.find_element_by_id('valid_thru_fill').text
  cvv = driver.find_element_by_id('cvv_fill').text
  kitno = driver.find_element_by_id('kit_number_fill').text
  balance = driver.find_element_by_id('balance').text
  print(cardno, expiry, cvv, kitno, balance)
  output_save()
  driver.find_element_by_xpath('//*[@id="doc_body"]/div[1]/header/nav/div/i').click()

def main_script():
    global balance, input_card_number, cardno, expiry, cvv, kitno, cards, n, cards_list
    driver.find_element_by_id("mobile_number").click()
    driver.find_element_by_id("mobile_number").clear()
    driver.find_element_by_id("mobile_number").send_keys(input_workbook_mobile_number[x])
    driver.find_element_by_xpath("//button[@type='submit']").click()
    driver.find_element_by_id("mpin").click()
    driver.find_element_by_id("mpin").clear()
    driver.find_element_by_id("mpin").send_keys("2021")
    driver.find_element_by_xpath("//form[@id='MPIN_form']/div[2]/button").click()
    time.sleep(0.5)
    cards_list = []
    elem = driver.find_element_by_xpath('//*[@id="gc_list"]')
    all_li = elem.find_elements_by_tag_name("li")
    for li in all_li:
      cards = li.get_attribute("id")
      cards_list.append(cards)
    print(len(cards_list))
    for n in range(0, len(cards_list)):
      demo()

    
    #c = driver.find_elements_by_tag_name("li")
    #print (c)
    #print(len(c))
    
    #b = driver.find_element_by_xpath("//li[@class='item gcl ACTIVE_f ']")
    #print(b.text)
    #a = driver.find_element_by_xpath("//li[@class='item gcl ACTIVE_f ']")
    #print (a)
    #for li in a:
      #print(a.get_attribute("id"))
    #print(a.get_attribute("id"))
    #cc_last_number()
    #driver.find_element_by_id(list_xxxx).click()
    time.sleep(3)
    #balance = driver.find_element_by_id('balance').text

    #driver.find_element_by_id('list_'+str(cc_last_number_12[x])).click()
    #time.sleep(2)
    '''
    #time.sleep(1)
    try:
        driver.find_element_by_xpath("//div[@id='toast-container']/div/div")
    except NoSuchElementException:
        driver.find_element_by_id("mpin").click()
        driver.find_element_by_id("mpin").clear()
        driver.find_element_by_id("mpin").send_keys("2244")
        driver.find_element_by_xpath("//form[@id='MPIN_form']/div[2]/button").click()
        time.sleep(200)
    else :
        driver.quit()
        
        driver.get("https://sr-giftcard.novopay.in/SR-Busines/Login")
        driver.find_element_by_id("mobile_number").click()
        driver.find_element_by_id("mobile_number").clear()
        driver.find_element_by_id("mobile_number").send_keys(input_workbook_mobile_number[x])
        driver.find_element_by_xpath("//button[@type='submit']").click()
        time.sleep(6)
    '''
    #print(driver.find_element_by_xpath("//span[text()='xxxx xxxx xxxx 2224']"))
    #//span[text()='thisisatest']
    #driver.find_element_by_xpath("//div[@id='View-Card']/span").text
    #driver.find_element_by_xpath("/html/body/div[1]/div[1]/div/section/div/div/div/div[2]/ul/li[1]/div[2]/span").text
    #//*[@id="View-Card"]/span
    #driver.find_elements_by_xpath("//span[@class='product-description card_number']")
    #print(driver.find_element_by_xpath("//span[@class='product-description card_number']"))
    #time.sleep(2)
    #driver.find_element_by_id("balance").click()

# get-output sheet to append output
output_sheet = path.exists("Output.xlsx")
if output_sheet == True :
  output_sheet_file_path = "Output.xlsx"
else :
  output_headers= ['Mobile Number', 'Card Number', 'Expiry', 'CVV', 'Kit Number', 'Current Balance']
  overall_output = Workbook()
  page = overall_output.active
  page.append(output_headers)
  overall_output.save(filename = 'Output.xlsx')
  output_sheet_file_path = "Output.xlsx"

def output_save():
  entry_list = [[ input_workbook_mobile_number[x], cardno, expiry, cvv, kitno, balance]]
  output_wb = load_workbook(output_sheet_file_path)
  page = output_wb.active
  for info in entry_list:
      page.append(info)
  output_wb.save(filename='Output.xlsx')

def resume_output():
  global output_cc_number
  global h
  output_load_wb_2 = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', usecols = 'A', dtype=int)
  output_load_wb_2.head()
  output_cc_number = output_load_wb_2['Mobile Number'].values.tolist()
  h = len(output_load_wb_2.index)

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")
caps = DesiredCapabilities().CHROME
#caps["pageLoadStrategy"] = "none"
#caps["pageLoadStrategy"] = "eager"
caps["pageLoadStrategy"] = "normal"
#driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
#driver.maximize_window()
try:
  resume_output()
except IndexError:
    for x in range (0 , number_of_rows):
        driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
        driver.maximize_window()
        start_link()
        main_script()
        driver.quit()
else:
    last_txncard = h
    for x in range (last_txncard , number_of_rows):
        driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
        driver.maximize_window()
        start_link()
        main_script()
        driver.quit()
